import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# Define the folder containing the Excel files
folder_path = "Sample logs"  # Folder path
output_file_name = "merged_log.xlsx"  # Desired merged file name

# Get all Excel files in the folder
if not os.path.exists(folder_path):
    print(f"The folder '{folder_path}' does not exist. Please check the path.")
    exit()

# Get all Excel files in the directory
excel_files = [file for file in os.listdir(folder_path) if file.endswith('.xlsx')]

# Check if there are any Excel files in the folder
if not excel_files:
    print(f"No Excel files found in the folder '{folder_path}'. Please add some files.")
    exit()

# Initialize an empty DataFrame to hold the merged data
merged_data = pd.DataFrame()

# Process each Excel file
for file in excel_files:
    file_path = os.path.join(folder_path, file)
    try:
        # Load all sheets from the Excel file
        excel_data = pd.read_excel(file_path, sheet_name=None)
    except Exception as e:
        print(f"Error reading file '{file}': {e}")
        continue

    for sheet_name, data in excel_data.items():
        # Add a column to identify the source file and sheet
        data['Source_File'] = file
        data['Source_Sheet'] = sheet_name
        # Append data to the merged DataFrame
        merged_data = pd.concat([merged_data, data], ignore_index=True)

# Save the merged data to an Excel file using pandas
temp_file_path = os.path.join(folder_path, output_file_name)
merged_data.to_excel(temp_file_path, index=False)

# Apply formatting using openpyxl
wb = load_workbook(temp_file_path)
ws = wb.active

# Increase row height and column width for better visibility
for row in ws.iter_rows():
    for cell in row:
        # Center the text horizontally and vertically
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # Set font size and style for readability
        cell.font = Font(size=14)  # Increase font size for better visibility

# Adjust the header row height and all columns for better visibility
ws.row_dimensions[1].height = 30  # Increase header row height

# Set column widths dynamically
for col in range(1, ws.max_column + 1):
    # Adjust the "Source_File" column width to be much wider
    if ws.cell(row=1, column=col).value == "Source_File":
        ws.column_dimensions[get_column_letter(col)].width = 50  # Make the Source_File column significantly wide
    else:
        ws.column_dimensions[get_column_letter(col)].width = 20  # Set the default width for other columns

# Save the styled merged file in the `sample logs` folder
wb.save(temp_file_path)

# Print confirmation that the file has been saved
print(f"Merged file saved in folder: {temp_file_path}")
